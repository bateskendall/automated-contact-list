import base64
import re
import os
import openpyxl
import pickle
import email
from openpyxl import Workbook
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

def get_credentials(account_folder):

    creds = None

    token_path = os.path.join(account_folder, 'token.pickle')
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)

    # If there are no (valid) credentials available, prompt the user to log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            credentials_path = os.path.join(account_folder, 'credentials.json')
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)

    return creds


def get_forwarded_emails(service):
    messages = []
    page_token = None

    while True:
        response = service.users().messages().list(
            userId='me', q='subject:FW OR subject:Fwd', maxResults=500, pageToken=page_token).execute()

        messages.extend(response['messages'])

        if 'nextPageToken' not in response:
            break

        page_token = response['nextPageToken']

    total_messages_fetched = len(messages)

    forwarded_emails = []

    for message in messages:
        msg = service.users().messages().get(userId='me', id=message['id'], format='full').execute()
        email_body = get_email_body(msg)
        from_lines = find_from_lines(email_body)
        message_id = msg['id']

        forwarded_emails.append({
            'email_body': email_body,
            'from_lines': from_lines,
            'message_id': message_id
        })

    return forwarded_emails, total_messages_fetched




def get_email_body(msg):
    if msg['payload'].get('mimeType') == 'text/plain':
        data = msg['payload'].get('body', {}).get('data')
        if data:
            return base64.urlsafe_b64decode(data).decode()

    message_parts = msg['payload'].get('parts')
    if message_parts:
        for part in message_parts:
            if part.get('mimeType') == 'text/plain':
                data = part['body'].get('data')
                if data:
                    return base64.urlsafe_b64decode(data).decode()
            elif part.get('mimeType') == 'multipart/alternative':
                subparts = part.get('parts')
                if subparts:
                    for subpart in subparts:
                        if subpart.get('mimeType') == 'text/plain':
                            data = subpart['body'].get('data')
                            if data:
                                return base64.urlsafe_b64decode(data).decode()
    return ''

def find_from_lines(email_body):
    from_pattern = r'(?i)From:[^\n]*'
    matches = re.findall(from_pattern, email_body, re.MULTILINE)
    return matches

def save_to_excel(forwarded_emails, output_path, total_messages_fetched):
    wb = openpyxl.Workbook()
    ws = wb.active

    row_num = 1
    for email in forwarded_emails:
        for from_line in email['from_lines']:
            name, email_address = extract_name_and_email(from_line)
            ws.cell(row=row_num, column=1, value=name)
            ws.cell(row=row_num, column=2, value=email_address)
            row_num += 1

    ws.cell(row=row_num, column=1, value=f'Total messages fetched: {total_messages_fetched}')
    
    wb.save(output_path)
    print(f"Data saved to {output_path}")

def extract_name_and_email(from_line):
    match = re.search(r'(?<=From: ).*?(?= <)', from_line)
    name = match.group(0) if match else ''
    
    match = re.search(r'(?<=<).*?(?=>)', from_line)
    email_address = match.group(0) if match else ''
    
    return name, email_address


def main():
    account_folders = ['account']  # Add or remove folders as needed
    output_path = 'forwarded_emails.xlsx'

    all_forwarded_emails = []

    for account_folder in account_folders:
        print(f"Processing emails for account: {account_folder}")
        creds = get_credentials(account_folder)
        service = build('gmail', 'v1', credentials=creds)

        forwarded_emails, total_messages_fetched = get_forwarded_emails(service)
        all_forwarded_emails.extend(forwarded_emails)

    save_to_excel(all_forwarded_emails, output_path, total_messages_fetched)
    print(f"Data saved to {output_path}")


if __name__ == '__main__':
    main()
